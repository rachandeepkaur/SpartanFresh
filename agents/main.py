"""FastAPI service exposing the Marathon Kitchen agent pipeline."""

from __future__ import annotations

from dotenv import load_dotenv

load_dotenv()  # must run before importing modules that read env vars at import time

import csv
import io

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import storage
from connectors import read_all
from freshness import estimate_freshness
from briefing import generate_briefs
from graph import run_pipeline
from prioritization import prioritize
from schemas import CanonicalEvent
from translation import translate, translate_upload_row

app = FastAPI(title="Marathon Kitchen Coordination Platform")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class IngestRequest(BaseModel):
    source_partner: str
    raw: dict


def _recompute_from_stored_events() -> dict:
    """Re-run freshness -> prioritize -> brief over whatever is currently
    in the store (used after seeding or after a live ingest)."""
    events = [CanonicalEvent(**e) for e in storage.list_events()]
    tags = [estimate_freshness(e) for e in events]
    ranked = prioritize(events, tags)
    briefs = generate_briefs(ranked)
    storage.save_briefs([b.model_dump(mode="json") for b in briefs])
    return {"event_count": len(events), "brief_count": len(briefs)}


@app.get("/health")
def health():
    return {"status": "ok", "using_supabase": storage.using_supabase()}


@app.post("/demo/seed")
def seed_demo_data():
    """Load the mock partner files, run the full pipeline, persist results.
    This is the 'aha' demo step: three different raw formats in, one
    unified inventory + two partner briefs out."""
    raw_records = read_all()
    if not raw_records:
        raise HTTPException(500, "No mock partner data found")

    result = run_pipeline(raw_records)
    event_dicts = [e.model_dump(mode="json") for e in result["events"]]
    for source_partner in {event["source_partner"] for event in event_dicts}:
        storage.replace_events_for_source(
            source_partner,
            [
                event
                for event in event_dicts
                if event["source_partner"] == source_partner
            ],
        )
    summary = _recompute_from_stored_events()

    return {
        "events_ingested": len(result["events"]),
        "by_source": {
            source: sum(1 for e in result["events"] if e.source_partner == source)
            for source in {s for s, _ in raw_records}
        },
        "briefs_generated": [b.partner_id for b in result["briefs"]],
        **summary,
    }


@app.post("/ingest")
def ingest(request: IngestRequest):
    """Translate and store a single live record, then refresh briefs."""
    event = translate(request.source_partner, request.raw)
    storage.insert_events([event.model_dump(mode="json")])
    summary = _recompute_from_stored_events()
    return {"event_id": event.id, **summary}


@app.post("/pipeline/refresh")
def refresh_pipeline():
    """Recompute freshness/priority/briefs from whatever is currently stored."""
    return _recompute_from_stored_events()


def _rows_from_csv_bytes(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    return [row for row in csv.DictReader(io.StringIO(text)) if any(row.values())]


def _rows_from_xlsx_bytes(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    result = []
    for row in rows:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
    return result


@app.post("/uploads")
async def upload_inventory_sheet(
    file: UploadFile = File(...),
    source_partner: str = Form(...),
):
    """A program manager attaches a partner's own spreadsheet (csv/xlsx) as
    a manual stand-in for a live connector. Each row is translated with the
    same canonical schema every other source lands in, so it flows through
    freshness/prioritization/briefing exactly like seeded or /ingest data."""
    source_partner = source_partner.strip()
    if not source_partner:
        raise HTTPException(400, "source_partner is required")

    filename = file.filename or "upload"
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    content = await file.read()

    if suffix == "csv":
        rows = _rows_from_csv_bytes(content)
    elif suffix in ("xlsx", "xlsm"):
        rows = _rows_from_xlsx_bytes(content)
    else:
        raise HTTPException(400, "Only .csv or .xlsx files are supported")

    if not rows:
        raise HTTPException(400, "No data rows found in the uploaded file")

    events: list[CanonicalEvent] = []
    errors: list[dict] = []
    for i, row in enumerate(rows, start=2):  # header is row 1
        try:
            events.append(translate_upload_row(source_partner, row))
        except Exception as exc:
            errors.append({"row": i, "error": str(exc)})

    inferred_sources = sorted({event.source_partner for event in events})
    if events:
        event_dicts = [event.model_dump(mode="json") for event in events]
        for inferred_source in inferred_sources:
            storage.replace_events_for_source(
                inferred_source,
                [
                    event
                    for event in event_dicts
                    if event["source_partner"] == inferred_source
                ],
            )
        # If the UI selection was only a fallback and the sheet identified a
        # different organization, remove the obsolete snapshot under the
        # incorrectly selected name.
        if source_partner not in inferred_sources:
            storage.replace_events_for_source(source_partner, [])
    summary = _recompute_from_stored_events()

    return {
        "filename": filename,
        "source_partner": source_partner,
        "source_partners": inferred_sources,
        "rows_read": len(rows),
        "events_ingested": len(events),
        "errors": errors,
        **summary,
    }


@app.get("/events")
def get_events():
    return storage.list_events()


@app.get("/briefs")
def get_briefs():
    return storage.list_briefs()


@app.get("/briefs/{partner_id}")
def get_brief(partner_id: str):
    briefs = {b["partner_id"]: b for b in storage.list_briefs()}
    if partner_id not in briefs:
        raise HTTPException(404, f"No brief for partner '{partner_id}'")
    return briefs[partner_id]
